from flask import Flask, jsonify, render_template, url_for, request, redirect, flash, session
from flask_login import LoginManager, login_user, logout_user, login_required
from werkzeug.security import generate_password_hash, check_password_hash
import flask
import psycopg2
import psycopg2.extras
import pandas as pd
from openpyxl.styles import Alignment
from fpdf import FPDF
from flask import send_file
from openpyxl import load_workbook


import re

from database.base import get_conection
conn = get_conection()

matricula = flask.Blueprint('matricula', __name__)

# Ruta para mostrar la página de nueva matrícula
@matricula.route('/templates/app/matriculas/nueva_matricula.html')
def nueva_matricula():
    if 'loggedin' in session:
        curc = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # Obtener los estudiantes para mostrar en el formulario
        sc = "SELECT id, nombres || ' ' || apellidos as nombre_completo FROM personas.personas WHERE nombre_rol = 'Estudiante'"
        curc.execute(sc)
        list_cpersonas = curc.fetchall()
        return render_template('app/matriculas/nueva_matricula.html', list_cpersonas=list_cpersonas)
    flash("Debes iniciar sesión para acceder a esta página.")
    return render_template('web/loginn.html')

# Ruta para buscar un estudiante por cédula
@matricula.route('/buscar_estudiante/<string:cedula>', methods=['GET'])
def buscar_estudiante(cedula):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    query = """
        SELECT id, nombres, apellidos, fecha_nacimiento, fecha_inscripcion, 
               nombre_cinturon, genero, nombre_rol, correo, telefono, direccion
        FROM personas.personas 
        WHERE id = %s AND nombre_rol = 'Estudiante'
    """
    cur.execute(query, (cedula,))
    estudiante = cur.fetchone()

    if estudiante:
        estudiante['fecha_nacimiento'] = estudiante['fecha_nacimiento'].strftime("%Y-%m-%d") if estudiante['fecha_nacimiento'] else None
        estudiante['fecha_inscripcion'] = estudiante['fecha_inscripcion'].strftime("%Y-%m-%d") if estudiante['fecha_inscripcion'] else None
        return jsonify({
            'id': estudiante['id'],
            'nombres': estudiante['nombres'],
            'apellidos': estudiante['apellidos'],
            'fecha_nacimiento': estudiante['fecha_nacimiento'],
            'fecha_inscripcion': estudiante['fecha_inscripcion'],
            'nombre_cinturon': estudiante['nombre_cinturon'],
            'genero': estudiante['genero'],
            'nombre_rol': estudiante['nombre_rol'],
            'correo': estudiante['correo'],
            'telefono': estudiante['telefono'],
            'direccion': estudiante['direccion']
        })
    else:
        return jsonify({'error': 'Estudiante no encontrado'})

# Ruta para registrar una nueva matrícula
@matricula.route('/registrar_matricula', methods=['POST'])
def registrar_matricula():
    if 'loggedin' in session:
        print("Datos enviados desde el formulario:", request.form)

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        id_persona = request.form.get('id_persona')
        fecha_matricula = request.form.get('fecha_matricula') or 'NOW()'

        print(f"id_persona: {id_persona}, fecha_matricula: {fecha_matricula}")

        # Validar si el estudiante existe
        cur.execute("SELECT * FROM personas.personas WHERE id = %s AND nombre_rol = 'Estudiante'", (id_persona,))
        estudiante = cur.fetchone()

        # if not estudiante:
        #     flash("El estudiante no existe o no es válido.")
        #     return redirect(url_for('matriculas.nueva_matricula'))

        # Verificar si ya existe una matrícula para el mismo id_personas
        cur.execute("SELECT * FROM personas.detalle_matricula WHERE id_personas = %s", (id_persona,))
        matricula_existente = cur.fetchone()

        if matricula_existente:
            flash("El estudiante ya está matriculado.")
            return redirect(url_for('matricula.nueva_matricula'))

        try:
            # Guardar matrícula en personas.detalle_matricula
            cur.execute("""
                INSERT INTO personas.detalle_matricula (id_personas, fecha_matricula)
                VALUES (%s, %s)
            """, (id_persona, fecha_matricula))
            conn.commit()
            flash("¡Matrícula registrada con éxito!")
        except Exception as e:
            conn.rollback()
            flash(f"Error al registrar la matrícula: {e}")
            print(f"Error al guardar en la base de datos: {e}")

        return redirect(url_for('matricula.ver_matriculas'))

    flash("Debes iniciar sesión para acceder a esta página.")
    return render_template('web/loginn.html')


# Ruta para listar las matrículas
@matricula.route('/templates/app/matriculas/matricula.html', methods=['POST', 'GET'])
def ver_matriculas():
    if 'loggedin' in session:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        if request.method == "POST" and 'buscar' in request.form and request.form['buscar'].strip():
            # Consulta para buscar por cédula
            search_term = '%' + request.form['buscar'].strip() + '%'
            s = """
                SELECT dm.id, dm.id_personas, p.nombres || ' ' || p.apellidos AS nombre_completo, 
                       dm.fecha_matricula, p.correo
                FROM personas.detalle_matricula dm
                INNER JOIN personas.personas p ON dm.id_personas = p.id
                WHERE dm.id_personas LIKE %s
            """
            cur.execute(s, (search_term,))
        else:
            # Mostrar todos los matriculados
            s = """
                SELECT dm.id, dm.id_personas, p.nombres || ' ' || p.apellidos AS nombre_completo, 
                       dm.fecha_matricula, p.correo
                FROM personas.detalle_matricula dm
                INNER JOIN personas.personas p ON dm.id_personas = p.id
            """
            cur.execute(s)

        matriculados = cur.fetchall()

        return render_template('app/matriculas/matricula.html', matriculados=matriculados)

    flash("Debes iniciar sesión para acceder a esta página.")
    return render_template('web/loginn.html')


# eliminar la matricula
@matricula.route('/delete_matricula/<string:id>', methods=['POST', 'GET'])
def delete_matricula(id):
    if 'loggedin' in session:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("DELETE FROM personas.detalle_matricula WHERE id = %s", (id,))
        conn.commit()
        print(id)
        flash('Matricula Eliminada Correctamente') 
        return redirect(url_for('matricula.ver_matriculas'))
    flash('Debes iniciar sesión para acceder a esta página.')
    return render_template('web/loginn.html')


# metodo para hacer un excel
from openpyxl import load_workbook

# Ruta para exportar las matrículas a Excel con ajuste de columnas
@matricula.route('/exportar_matriculas_excel', methods=['GET'])
def exportar_matriculas_excel():
    if 'loggedin' in session:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Consulta para obtener todas las matrículas
            cur.execute("""
                SELECT dm.id_personas, p.nombres || ' ' || p.apellidos AS nombre_completo, 
                       dm.fecha_matricula, p.correo
                FROM personas.detalle_matricula dm
                INNER JOIN personas.personas p ON dm.id_personas = p.id
            """)
            matriculas = cur.fetchall()

            # Crear un DataFrame de pandas con formato
            df = pd.DataFrame(matriculas, columns=['Cédula', 'Nombre Completo', 'Fecha Matrícula', 'Correo'])

            # Formatear la columna de fechas
            if 'Fecha Matrícula' in df.columns:
                df['Fecha Matrícula'] = pd.to_datetime(df['Fecha Matrícula']).dt.strftime('%Y-%m-%d')

            # Exportar a un archivo Excel
            excel_file = 'matriculas.xlsx'
            df.to_excel(excel_file, index=False, engine='openpyxl')

            # Ajustar el ancho de las columnas y centrar el texto
            wb = load_workbook(excel_file)
            ws = wb.active

            # Ajustar el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                    # Centrar el contenido de cada celda
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[col_letter].width = max_length + 2  # Margen adicional
            
            # Guardar el archivo
            wb.save(excel_file)

            # Devolver el archivo como descarga
            return flask.send_file(excel_file, as_attachment=True)
        except Exception as e:
            print(f"Error al exportar matrículas a Excel: {e}")
            flash("Error al exportar las matrículas.")
            return redirect(url_for('matricula.ver_matriculas'))
    else:
        flash("Debes iniciar sesión para acceder a esta página.")
        return render_template('web/loginn.html')
    
# generador de pdf para matricula
@matricula.route('/generar_matricula_pdf/<string:id>', methods=['GET'])
def generar_matricula_pdf(id):
    if 'loggedin' in session:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Obtener información del estudiante y la matrícula
            cur.execute("""
                SELECT dm.id_personas, p.nombres || ' ' || p.apellidos AS nombre_completo,
                       dm.fecha_matricula, p.fecha_nacimiento, p.correo, p.telefono
                FROM personas.detalle_matricula dm
                INNER JOIN personas.personas p ON dm.id_personas = p.id
                WHERE dm.id = %s
            """, (id,))
            matricula = cur.fetchone()

            if not matricula:
                flash("No se encontró la matrícula.")
                return redirect(url_for('matricula.ver_matriculas'))

            # Crear el PDF
            pdf = FPDF()
            pdf.add_page()

            # Agregar el logo
            pdf.image('static/img/portada.jpeg', x=10, y=8, w=25)  # Ajusta la ruta y tamaño del logo

            # Encabezado
            pdf.set_font("Arial", style="B", size=14)
            pdf.cell(200, 10, txt="Academia de Artes Marciales Team Ray", ln=True, align='C')
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt=f"Fecha de emisión: {pd.Timestamp.now().strftime('%Y-%m-%d')}", ln=True, align='C')
            pdf.ln(20)

            # Información del Estudiante
            pdf.set_font("Arial", style="B", size=12)
            pdf.cell(200, 10, txt="Información del Estudiante", ln=True)
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt=f"- Nombre: {matricula['nombre_completo']}", ln=True)
            pdf.cell(200, 10, txt=f"- Cédula: {matricula['id_personas']}", ln=True)
            pdf.cell(200, 10, txt=f"- Fecha de Nacimiento: {matricula['fecha_nacimiento']}", ln=True)
            pdf.cell(200, 10, txt=f"- Correo Electrónico: {matricula['correo']}", ln=True)
            pdf.cell(200, 10, txt=f"- Teléfono: {matricula['telefono']}", ln=True)
            pdf.ln(10)

            # Detalles de la Matrícula
            pdf.set_font("Arial", style="B", size=12)
            pdf.cell(200, 10, txt="Detalles de la Matrícula", ln=True)
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt=f"- Fecha de Matrícula: {matricula['fecha_matricula']}", ln=True)
            pdf.cell(200, 10, txt=f"- Nivel/Cinturón: Cinturón Blanco", ln=True)  # Cambiar según sea necesario
            pdf.ln(20)

            # Espacio para firmas lado a lado
            pdf.set_font("Arial", style="B", size=12)
            pdf.cell(90, 10, txt="Firma del Estudiante:", ln=0, align='L')  # En el lado izquierdo
            pdf.cell(90, 10, txt="Firma del Instructor:", ln=1, align='R')  # En el lado derecho
            pdf.cell(90, 35, txt="___________________________", ln=0, align='L')  # Línea para firma izquierda
            pdf.cell(90, 35, txt="___________________________", ln=1, align='R')  # Línea para firma derecha
            pdf.ln(20)

            # Colocar el pie de página de forma precisa
            pdf.set_y(-55)  # Ajustar posición al final de la página
            pdf.set_font("Arial", size=10)
            pdf.cell(200, 10, txt="Dirección: Av. Principal 123, Quito, Ecuador", ln=True, align='C')
            pdf.cell(200, 10, txt="Teléfono: +593 123456789 | Correo: contacto@academia.com", ln=True, align='C')
            pdf.cell(200, 10, txt="Página Web: www.academia.com", ln=True, align='C')

            # Guardar y enviar el PDF
            pdf_file = f"matricula_{id}.pdf"
            pdf.output(pdf_file)
            return send_file(pdf_file, as_attachment=True)

        except Exception as e:
            print(f"Error al generar el PDF: {e}")
            flash("Error al generar el PDF.")
            return redirect(url_for('matricula.ver_matriculas'))
    else:
        flash("Debes iniciar sesión para acceder a esta página.")
        return render_template('web/loginn.html')



